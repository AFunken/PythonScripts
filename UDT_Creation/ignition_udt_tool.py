"""
Ignition UDT Tool
=================

One tool, two commands:

  1) template  -> generate a rich, self-documenting Excel workbook with
                  dropdowns (data validation) for every field whose allowable
                  values are defined by Inductive Automation.

  2) build     -> read a filled-out workbook (or the old single-sheet format)
                  and emit an Ignition UDT definition JSON ready to import.

Usage
-----
  # 1. Create a blank template to fill out
  python ignition_udt_tool.py template MyTemplate.xlsx

  # 2. Fill it out in Excel, then convert to JSON
  python ignition_udt_tool.py build MyTemplate.xlsx output.json

  # (Backwards compatible) build straight from an old flat sheet:
  python ignition_udt_tool.py build STD_DI.xlsx STD_DI.json --udt-name STD_DI

Requires: pandas, openpyxl
"""

import os
import re
import sys
import json
import argparse

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.comments import Comment


# ---------------------------------------------------------------------------
# Allowable values defined by Inductive Automation (Ignition 8.1 tag JSON).
# These feed the dropdowns AND validate the build step.
# ---------------------------------------------------------------------------
ALLOWED = {
    # tag data types
    "dataType": [
        "Boolean", "Int1", "Int2", "Int4", "Int8", "Float4", "Float8",
        "String", "DateTime", "BooleanArray", "Int1Array", "Int2Array",
        "Int4Array", "Int8Array", "Float4Array", "Float8Array", "StringArray",
        "DateTimeArray", "DataSet", "ByteArray", "Document",
    ],
    # NOTE: the tag editor labels a SQL query tag "Query", but the JSON key
    # value is "db". These are the actual JSON strings.
    "valueSource": ["opc", "memory", "expr", "db", "reference", "derived"],
    "tagType": ["AtomicTag", "Folder", "UdtType", "UdtInstance"],
    "sampleMode": ["OnChange", "Periodic", "TagGroup"],
    "historicalDeadbandStyle": ["Auto", "Analog_Compressed", "Discrete"],
    "sampleRateUnits": ["MS", "SEC", "MIN", "HOUR", "DAY", "WEEK", "MONTH", "YEAR"],
    "deadbandMode": ["Absolute", "Percent", "Off"],
    "alarmMode": [
        "WhenTrue", "WhenFalse", "Equality", "AboveValue", "BelowValue",
        "BetweenValues", "OutsideValues", "OutOfRange", "AnyChange",
        "BadQuality", "OnCondition",
    ],
    "priority": ["Diagnostic", "Low", "Medium", "High", "Critical"],
    "boolean": ["TRUE", "FALSE"],
    "parameterType": ["String", "Integer", "Float", "Boolean"],
}

# Providers / tag groups have no fixed enum (they are project-specific), so we
# seed these editable lists with common defaults the user can extend.
SEED_PROVIDERS = ["", "Historian", "Sensor_Data", "Edge Historian"]
SEED_TAGGROUPS = ["", "default", "Default Historical", "OneSecond", "OneMinute"]

# Columns on the Tags sheet, in fill-out order. (header, comment)
TAG_COLUMNS = [
    ("name",                     "REQUIRED. Member tag / parameter name (e.g. Inp_PVraw)."),
    ("tagType",                  "AtomicTag (default), Folder, UdtInstance, UdtType."),
    ("dataType",                 "Ignition data type. Dropdown."),
    ("valueSource",              "opc | memory | expr | db | reference | derived. Dropdown."),
    ("opcItemPath",              "OPC address. Reference UDT params with braces, e.g. ns=1;s=[{Device}]HRI352"),
    ("opcServer",                "OPC server. Bind to a param with {Server}, or type a literal name. Blank = default."),
    ("expression",               "Expression text (only for valueSource=expr)."),
    ("documentation",            "Built-in Ignition documentation field."),
    ("tooltip",                  "Short hover tooltip."),
    ("shortDescription",         "Metadata custom property -> Metadata.shortDescription (always emitted)."),
    ("longDescription",          "Metadata custom property -> Metadata.longDescription (always emitted)."),
    ("engUnit",                  "Engineering unit, e.g. V, A, Hz, degF."),
    ("engLow",                   "Engineering range low (numeric)."),
    ("engHigh",                  "Engineering range high (numeric)."),
    ("formatString",             "Display format, e.g. #,##0.##"),
    ("tagGroup",                 "Tag group / scan class. Editable dropdown."),
    ("deadband",                 "Value deadband (numeric)."),
    ("deadbandMode",             "Absolute | Percent | Off. Dropdown."),
    ("historyEnabled",           "TRUE/FALSE. Turns on tag historian for this tag."),
    ("historyProvider",          "Tag Historian provider name. Editable dropdown."),
    ("sampleMode",               "OnChange | Periodic | TagGroup. Dropdown."),
    ("historySampleRate",        "History sample rate (numeric, used when sampleMode=Periodic)."),
    ("historySampleRateUnits",   "MS | SEC | MIN | HOUR | DAY | WEEK | MONTH | YEAR. Dropdown."),
    ("historicalDeadband",       "Historical deadband (numeric)."),
    ("historicalDeadbandStyle",  "Auto | Analog_Compressed | Discrete. Dropdown."),
    ("historicalDeadbandMode",   "Absolute | Percent | Off. Dropdown."),
]

# Any of the value columns below accept THREE forms:
#   * a literal            -> 528   |  High  |  TRUE
#   * a UDT parameter      -> {VbcEnable}          (whole cell = one {Param})
#   * an expression        -> expr:toInt({VbcPriority})   (prefix with 'expr:')
# A cell that contains {braces} plus other text is treated as an expression too.
ALARM_COLUMNS = [
    ("tagName",              "REQUIRED. Must match a name on the Tags sheet."),
    ("alarmName",            "REQUIRED. Alarm name, e.g. 'Alarm' or 'High Alarm'."),
    ("mode",                 "Alarm condition. Dropdown."),
    ("setpointA",            "Primary setpoint. Number, {Param}, or expr:... For Boolean Equality use 1/0."),
    ("setpointB",            "Second setpoint (BetweenValues/OutsideValues). Number, {Param}, or expr:..."),
    ("label",                "Display label. Text, {Param}, or expr:{Description} + ' Phase AB Alarm'"),
    ("priority",             "Diagnostic|Low|Medium|High|Critical, OR {Param}, OR expr:toInt({VbcPriority})"),
    ("enabled",              "TRUE/FALSE, OR bind to a param e.g. {VbcEnable}"),
    ("ackNotesReqd",         "TRUE/FALSE. Require notes when acknowledging."),
    ("timeOnDelaySeconds",   "Delay before the alarm becomes active (numeric)."),
    ("timeOffDelaySeconds",  "Delay before the alarm clears (numeric)."),
    ("displayPath",          "Optional friendly path shown in alarm displays."),
]

# States feed Metadata.states -> [{label, value}, ...]. Multiple rows per tag.
STATE_COLUMNS = [
    ("tagName", "REQUIRED. Must match a name on the Tags sheet."),
    ("label",   "State label, e.g. Normal / Alarm / Running / Stopped."),
    ("value",   "State value: TRUE/FALSE, an integer (0,1,2...), or text. Type auto-detected."),
]

# ---------- styling helpers ----------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(italic=True, size=10, color="595959")
LABEL_FONT = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws, headers_with_comments, row=1):
    for c, (header, comment) in enumerate(headers_with_comments, start=1):
        cell = ws.cell(row=row, column=c, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        if comment:
            cell.comment = Comment(comment, "Ignition UDT Tool")
        ws.column_dimensions[get_column_letter(c)].width = max(14, min(30, len(header) + 6))
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _add_list_validation(ws, name, col_letter, first_row=2, last_row=500,
                         allow_blank=True, strict=True):
    # strict=False keeps the convenient dropdown but still ALLOWS free text
    # (needed for fields that may hold a {Param} or expr: binding instead).
    dv = DataValidation(type="list", formula1=f"={name}", allowBlank=allow_blank,
                        showErrorMessage=strict)
    if strict:
        dv.error = "Pick a value from the dropdown list."
        dv.errorTitle = "Invalid entry"
    dv.prompt = "Choose from the list (or type a {Param} / expr: binding)" \
        if not strict else "Choose from the list"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


# ===========================================================================
# TEMPLATE GENERATION
# ===========================================================================
def _new_workbook_skeleton():
    """Build the full template workbook (all sheets, dropdowns, named ranges,
    UDT layout, headers) with NO data rows. Returns (wb, refs)."""
    wb = Workbook()

    # ---- Instructions sheet ----
    ws_help = wb.active
    ws_help.title = "Instructions"
    ws_help.sheet_view.showGridLines = False
    ws_help["A1"] = "Ignition UDT Definition Builder"
    ws_help["A1"].font = TITLE_FONT
    steps = [
        "",
        "How to use this workbook:",
        "  1. Fill out the 'UDT' sheet: the UDT name, optional parent type, and any UDT Parameters (e.g. Device).",
        "  2. Fill out the 'Tags' sheet: one row per member tag. Cells with a dropdown arrow are validated.",
        "  3. (Optional) Fill out the 'Alarms' sheet: one row per alarm. 'tagName' must match a name on 'Tags'.",
        "  4. (Optional) Fill out the 'States' sheet: one row per state. 'tagName' must match a name on 'Tags'.",
        "  5. Save, then run:  python ignition_udt_tool.py build <thisfile>.xlsx output.json",
        "  6. In Ignition: Tag Browser > Import icon > select output.json.",
        "",
        "Notes:",
        "  - Every tag ALWAYS gets a 'Metadata' custom property with shortDescription + longDescription.",
        "    Add states for a tag on the 'States' sheet and they land under Metadata.states.",
        "  - Only 'name' is required per tag row. Leave any field blank to use the Ignition default.",
        "  - The 'Lists' sheet holds all allowable values. You may extend 'historyProvider' and 'tagGroup' there.",
        "  - Define UDT parameters on the 'UDT' sheet, then reference them with braces anywhere:",
        "      opcItemPath: ns=1;s=[{Device}]HRI352     opcServer: {Server}",
        "  - On the 'Alarms' sheet a value cell can be a literal, a parameter, or an expression:",
        "      enabled: {AlarmEnable}       (binds to a UDT parameter)",
        "      priority: expr:toInt({AlarmPriority})    label: expr:{Description} + ' Alarm'",
        "  - 'valueSource' uses the JSON value 'db' for a SQL Query tag (the editor labels it 'Query').",
    ]
    for i, line in enumerate(steps, start=2):
        ws_help.cell(row=i, column=1, value=line).font = SUB_FONT if line.startswith("  ") else LABEL_FONT
    ws_help.column_dimensions["A"].width = 110

    # ---- Lists sheet (feeds all dropdowns via named ranges) ----
    ws_lists = wb.create_sheet("Lists")
    ws_lists.sheet_view.showGridLines = False
    list_specs = [
        ("dataType",                "nDataType",         ALLOWED["dataType"]),
        ("valueSource",             "nValueSource",      ALLOWED["valueSource"]),
        ("tagType",                 "nTagType",          ALLOWED["tagType"]),
        ("sampleMode",              "nSampleMode",       ALLOWED["sampleMode"]),
        ("historicalDeadbandStyle", "nDeadbandStyle",    ALLOWED["historicalDeadbandStyle"]),
        ("sampleRateUnits",         "nSampleRateUnits",  ALLOWED["sampleRateUnits"]),
        ("deadbandMode",            "nDeadbandMode",     ALLOWED["deadbandMode"]),
        ("alarmMode",               "nAlarmMode",        ALLOWED["alarmMode"]),
        ("priority",                "nPriority",         ALLOWED["priority"]),
        ("boolean",                 "nBoolean",          ALLOWED["boolean"]),
        ("parameterType",           "nParameterType",    ALLOWED["parameterType"]),
        ("historyProvider*",        "nProvider",         SEED_PROVIDERS),   # * = editable/extendable
        ("tagGroup*",               "nTagGroup",         SEED_TAGGROUPS),
    ]
    EXTENDABLE_ROWS = 200  # allow the two editable lists to grow
    for col_idx, (label, name, values) in enumerate(list_specs, start=1):
        letter = get_column_letter(col_idx)
        hdr = ws_lists.cell(row=1, column=col_idx, value=label)
        hdr.font = HEADER_FONT
        hdr.fill = HEADER_FILL
        hdr.alignment = Alignment(horizontal="center")
        ws_lists.column_dimensions[letter].width = 22
        for r, v in enumerate(values, start=2):
            ws_lists.cell(row=r, column=col_idx, value=v)
        # editable lists get a generous range so users can append entries
        last = (1 + EXTENDABLE_ROWS) if label.endswith("*") else (1 + len(values))
        ref = f"Lists!${letter}$2:${letter}${last}"
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    # ---- UDT sheet (top-level info + parameters) ----
    ws_udt = wb.create_sheet("UDT")
    ws_udt.sheet_view.showGridLines = False
    ws_udt["A1"] = "UDT Definition"
    ws_udt["A1"].font = TITLE_FONT
    info_rows = [
        ("UDT Name",            "REQUIRED. e.g. STD_DI"),
        ("Parent Type (typeId)", "Optional. Name/path of a parent UDT to inherit from."),
        ("Default Tag Group",   "Optional. Applied to members that leave tagGroup blank."),
        ("Type Color (optional)", "Optional. Ignition tree color as an integer, e.g. -16777088."),
    ]
    for i, (label, hint) in enumerate(info_rows, start=3):
        lc = ws_udt.cell(row=i, column=1, value=label)
        lc.font = LABEL_FONT
        vc = ws_udt.cell(row=i, column=2)
        vc.border = BORDER
        vc.fill = PatternFill("solid", fgColor="FFF2CC")
        ws_udt.cell(row=i, column=3, value=hint).font = SUB_FONT
    # Parameters table
    ws_udt["A7"] = "UDT Parameters"
    ws_udt["A7"].font = Font(bold=True, size=12, color="1F4E78")
    param_hdr = [("Parameter Name", "e.g. Device"),
                 ("Type", "String | Integer | Float | Boolean"),
                 ("Value", "Default value for this parameter")]
    for c, (h, cm) in enumerate(param_hdr, start=1):
        cell = ws_udt.cell(row=8, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.comment = Comment(cm, "Ignition UDT Tool")
        cell.border = BORDER
    ws_udt.column_dimensions["A"].width = 22
    ws_udt.column_dimensions["B"].width = 16
    ws_udt.column_dimensions["C"].width = 40
    # validate parameter Type column
    dv_ptype = DataValidation(type="list", formula1="=nParameterType", allowBlank=True)
    ws_udt.add_data_validation(dv_ptype)
    dv_ptype.add("B9:B100")
    for r in range(9, 101):
        for c in range(1, 4):
            ws_udt.cell(row=r, column=c).border = BORDER

    # ---- Tags sheet ----
    ws_tags = wb.create_sheet("Tags")
    _style_header_row(ws_tags, TAG_COLUMNS, row=1)
    tag_col = {h: i + 1 for i, (h, _) in enumerate(TAG_COLUMNS)}

    def tcol(name):
        return get_column_letter(tag_col[name])

    _add_list_validation(ws_tags, "nDataType",        tcol("dataType"))
    _add_list_validation(ws_tags, "nValueSource",     tcol("valueSource"))
    _add_list_validation(ws_tags, "nTagType",         tcol("tagType"))
    _add_list_validation(ws_tags, "nSampleMode",      tcol("sampleMode"))
    _add_list_validation(ws_tags, "nSampleRateUnits", tcol("historySampleRateUnits"))
    _add_list_validation(ws_tags, "nDeadbandStyle",   tcol("historicalDeadbandStyle"))
    _add_list_validation(ws_tags, "nDeadbandMode",    tcol("deadbandMode"))
    _add_list_validation(ws_tags, "nDeadbandMode",    tcol("historicalDeadbandMode"))
    _add_list_validation(ws_tags, "nBoolean",         tcol("historyEnabled"))
    _add_list_validation(ws_tags, "nProvider",        tcol("historyProvider"))
    _add_list_validation(ws_tags, "nTagGroup",        tcol("tagGroup"))

    # ---- Alarms sheet ----
    ws_al = wb.create_sheet("Alarms")
    _style_header_row(ws_al, ALARM_COLUMNS, row=1)
    al_col = {h: i + 1 for i, (h, _) in enumerate(ALARM_COLUMNS)}
    _add_list_validation(ws_al, "nAlarmMode", get_column_letter(al_col["mode"]))
    # priority & enabled may hold a {Param}/expr: binding -> non-blocking dropdown
    _add_list_validation(ws_al, "nPriority", get_column_letter(al_col["priority"]), strict=False)
    _add_list_validation(ws_al, "nBoolean", get_column_letter(al_col["enabled"]), strict=False)
    _add_list_validation(ws_al, "nBoolean", get_column_letter(al_col["ackNotesReqd"]))

    # ---- States sheet (feeds Metadata.states) ----
    ws_st = wb.create_sheet("States")
    _style_header_row(ws_st, STATE_COLUMNS, row=1)
    st_col = {h: i + 1 for i, (h, _) in enumerate(STATE_COLUMNS)}

    return wb, {
        "ws_udt": ws_udt, "ws_tags": ws_tags, "ws_al": ws_al, "ws_st": ws_st,
        "tag_col": tag_col, "al_col": al_col, "st_col": st_col,
        "param_start_row": 9,
    }


def _finalize_and_save(wb, output_file):
    wb.move_sheet("Instructions", -wb.sheetnames.index("Instructions"))
    out_dir = os.path.dirname(os.path.abspath(output_file))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_file)


def make_template(output_file):
    """Generate a blank, self-documenting template pre-filled with examples."""
    wb, R = _new_workbook_skeleton()
    ws_udt, ws_tags, ws_al, ws_st = R["ws_udt"], R["ws_tags"], R["ws_al"], R["ws_st"]
    tag_col, al_col, st_col = R["tag_col"], R["al_col"], R["st_col"]
    pstart = R["param_start_row"]

    ws_udt["B3"] = "MyUDT"
    example_params = [
        ("Device", "String", None),
        ("Server", "String", "Ignition OPC UA Server"),
        ("Description", "String", None),
        ("AlarmEnable", "String", "True"),
        ("AlarmPriority", "Integer", 2),
    ]
    for i, (pn, pt, pv) in enumerate(example_params):
        ws_udt.cell(row=pstart + i, column=1, value=pn)
        ws_udt.cell(row=pstart + i, column=2, value=pt)
        if pv is not None:
            ws_udt.cell(row=pstart + i, column=3, value=pv)

    examples = [
        {"name": "Inp_PVraw", "tagType": "AtomicTag", "dataType": "Boolean",
         "valueSource": "opc", "opcItemPath": "ns=1;s=[{Device}]HRI350",
         "opcServer": "{Server}",
         "documentation": "Physical digital input signal (raw state).",
         "shortDescription": "Raw Input",
         "longDescription": "Physical digital input signal (raw state before inversion or debounce).",
         "historyEnabled": "FALSE"},
        {"name": "Out_PVActv", "tagType": "AtomicTag", "dataType": "Boolean",
         "valueSource": "opc", "opcItemPath": "ns=1;s=[{Device}]HRI352",
         "opcServer": "{Server}",
         "documentation": "Final active process value after debounce.",
         "shortDescription": "Active PV",
         "longDescription": "Final active process value after simulation, inversion, and debounce.",
         "historyEnabled": "TRUE", "historyProvider": "Historian",
         "sampleMode": "OnChange", "historySampleRate": 5,
         "historySampleRateUnits": "MIN", "historicalDeadbandStyle": "Discrete"},
        {"name": "TVSS_ALM", "tagType": "AtomicTag", "dataType": "Boolean",
         "valueSource": "opc", "opcItemPath": "ns=1;s=[{Device}]C1",
         "opcServer": "{Server}",
         "documentation": "Surge Protection Device Alarm",
         "shortDescription": "Surge Protection Device Alarm",
         "longDescription": "Digital status of the surge protection device; see States for meaning."},
    ]
    for r, ex in enumerate(examples, start=2):
        for h, v in ex.items():
            ws_tags.cell(row=r, column=tag_col[h], value=v)

    ws_al.cell(row=2, column=al_col["tagName"], value="Out_PVActv")
    ws_al.cell(row=2, column=al_col["alarmName"], value="Alarm")
    ws_al.cell(row=2, column=al_col["mode"], value="Equality")
    ws_al.cell(row=2, column=al_col["setpointA"], value=1)
    ws_al.cell(row=2, column=al_col["label"], value="expr:{Description} + ' Active Alarm'")
    ws_al.cell(row=2, column=al_col["priority"], value="expr:toInt({AlarmPriority})")
    ws_al.cell(row=2, column=al_col["enabled"], value="{AlarmEnable}")
    ws_al.cell(row=2, column=al_col["ackNotesReqd"], value="TRUE")

    for r, (label, val) in enumerate([("Normal", "FALSE"), ("Alarm", "TRUE")], start=2):
        ws_st.cell(row=r, column=st_col["tagName"], value="TVSS_ALM")
        ws_st.cell(row=r, column=st_col["label"], value=label)
        ws_st.cell(row=r, column=st_col["value"], value=val)

    _finalize_and_save(wb, output_file)
    print(f"[OK] Template written: {output_file}")
    print("     Sheets: Instructions, UDT, Tags, Alarms, States, Lists")
    print("     Fill it out, then run:  python ignition_udt_tool.py build "
          f"\"{output_file}\" output.json")


# ===========================================================================
# IMPORT  (Ignition UDT JSON -> filled template workbook)
# ===========================================================================
def _unbind(v):
    """Reverse a JSON value/binding object into a spreadsheet cell string."""
    if isinstance(v, dict):
        bt = v.get("bindType")
        if bt == "parameter":
            return v.get("binding", "")
        if bt == "UDTParameter":
            return v.get("value", "")
        if bt == "Expression":
            return "expr:" + str(v.get("value", ""))
        return v.get("binding", v.get("value", ""))
    return v


def _state_cell(v):
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return v


def _write_tag_row(ws, tag_col, row, tag):
    meta = tag.get("Metadata") or {}

    def put(colname, val):
        if val is None or val == "":
            return
        ws.cell(row=row, column=tag_col[colname], value=val)

    put("name", tag.get("name"))
    put("tagType", tag.get("tagType"))
    put("dataType", tag.get("dataType"))
    put("valueSource", tag.get("valueSource"))
    put("opcItemPath", _unbind(tag.get("opcItemPath")))
    put("opcServer", _unbind(tag.get("opcServer")))
    put("expression", tag.get("expression"))
    put("documentation", tag.get("documentation"))
    put("tooltip", tag.get("tooltip"))
    put("shortDescription", meta.get("shortDescription"))
    put("longDescription", meta.get("longDescription"))
    put("engUnit", tag.get("engUnit"))
    put("engLow", tag.get("engLow"))
    put("engHigh", tag.get("engHigh"))
    put("formatString", tag.get("formatString"))
    put("tagGroup", tag.get("tagGroup"))
    put("deadband", tag.get("deadband"))
    put("deadbandMode", tag.get("deadbandMode"))
    if tag.get("historyEnabled") is not None:
        put("historyEnabled", "TRUE" if tag["historyEnabled"] else "FALSE")
    put("historyProvider", tag.get("historyProvider"))
    put("sampleMode", tag.get("sampleMode"))
    put("historySampleRate", tag.get("historySampleRate"))
    put("historySampleRateUnits", tag.get("historySampleRateUnits"))
    put("historicalDeadband", tag.get("historicalDeadband"))
    put("historicalDeadbandStyle", tag.get("historicalDeadbandStyle"))
    put("historicalDeadbandMode", tag.get("historicalDeadbandMode"))


def _write_alarm_row(ws, al_col, row, tagname, al):
    def put(colname, val):
        if val is None or val == "":
            return
        ws.cell(row=row, column=al_col[colname], value=val)

    put("tagName", tagname)
    put("alarmName", al.get("name"))
    put("mode", al.get("mode"))
    put("setpointA", _unbind(al.get("setpointA")))
    put("setpointB", _unbind(al.get("setpointB")))
    put("label", _unbind(al.get("label")))
    put("priority", _unbind(al.get("priority")))
    en = al.get("enabled")
    if isinstance(en, dict):
        put("enabled", _unbind(en))
    elif en is not None:
        put("enabled", "TRUE" if en else "FALSE")
    if al.get("ackNotesReqd") is not None:
        put("ackNotesReqd", "TRUE" if al["ackNotesReqd"] else "FALSE")
    put("timeOnDelaySeconds", al.get("timeOnDelaySeconds"))
    put("timeOffDelaySeconds", al.get("timeOffDelaySeconds"))
    put("displayPath", _unbind(al.get("displayPath")))


def import_json(json_file, output_file):
    """Read an Ignition UDT JSON and write it back into a fillable template."""
    with open(json_file, encoding="utf-8") as f:
        data = json.load(f)

    wb, R = _new_workbook_skeleton()
    ws_udt, ws_tags, ws_al, ws_st = R["ws_udt"], R["ws_tags"], R["ws_al"], R["ws_st"]
    tag_col, al_col, st_col = R["tag_col"], R["al_col"], R["st_col"]
    pstart = R["param_start_row"]

    ws_udt["B3"] = data.get("name", "")
    if data.get("typeId"):
        ws_udt["B4"] = data["typeId"]
    if data.get("typeColor") is not None:
        ws_udt["B6"] = data["typeColor"]

    r = pstart
    for pname, pdef in (data.get("parameters") or {}).items():
        pdef = pdef or {}
        ws_udt.cell(row=r, column=1, value=pname)
        ws_udt.cell(row=r, column=2, value=pdef.get("dataType", "String"))
        if "value" in pdef:
            ws_udt.cell(row=r, column=3, value=pdef["value"])
        r += 1

    trow = arow = srow = 2
    n_alarms = n_states = 0
    unmapped = set()
    known = set(tag_col) | {"Metadata", "alarms", "opcItemPath", "opcServer"}
    for tag in data.get("tags", []):
        _write_tag_row(ws_tags, tag_col, trow, tag)
        trow += 1
        name = tag.get("name", "")
        unmapped |= {k for k in tag if k not in known}
        for al in (tag.get("alarms") or []):
            _write_alarm_row(ws_al, al_col, arow, name, al)
            arow += 1
            n_alarms += 1
        for st in ((tag.get("Metadata") or {}).get("states") or []):
            ws_st.cell(row=srow, column=st_col["tagName"], value=name)
            ws_st.cell(row=srow, column=st_col["label"], value=st.get("label", ""))
            ws_st.cell(row=srow, column=st_col["value"], value=_state_cell(st.get("value")))
            srow += 1
            n_states += 1

    _finalize_and_save(wb, output_file)
    print(f"[OK] Imported {json_file}")
    print(f"     -> {output_file}")
    print(f"     UDT Name : {data.get('name', '')}")
    print(f"     Tags     : {len(data.get('tags', []))}")
    if n_alarms:
        print(f"     Alarms   : {n_alarms}")
    if n_states:
        print(f"     States   : {n_states}")
    if unmapped:
        print("     [note] tag fields with no template column were left out: "
              + ", ".join(sorted(unmapped)))


# ===========================================================================
# BUILD (workbook -> Ignition UDT JSON)
# ===========================================================================
def _s(v):
    """Clean string or None."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    return s if s else None


def _num(v, as_int=False):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return int(round(float(v))) if as_int else float(v)
    except (ValueError, TypeError):
        return None


def _bool(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("true", "1", "yes", "y")


def _read_udt_sheet(path):
    """Return (udt_name, parent, default_group, type_color, parameters_dict)."""
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return None, None, None, None, {}
    if "UDT" not in wb.sheetnames:
        return None, None, None, None, {}
    ws = wb["UDT"]
    name = _s(ws["B3"].value)
    parent = _s(ws["B4"].value)
    default_group = _s(ws["B5"].value)
    type_color = _num(ws["B6"].value, as_int=True)
    params = {}
    type_map = {"string": "String", "integer": "Integer", "int": "Integer",
                "float": "Float", "boolean": "Boolean"}
    r = 9
    while True:
        pname = _s(ws.cell(row=r, column=1).value)
        if not pname:
            if r > 9 and not _s(ws.cell(row=r + 1, column=1).value):
                break
            if r > 200:
                break
            r += 1
            continue
        ptype_raw = (_s(ws.cell(row=r, column=2).value) or "String")
        ptype = type_map.get(ptype_raw.lower(), "String")
        pval = ws.cell(row=r, column=3).value
        if ptype == "Integer":
            pval = _num(pval, as_int=True)
        elif ptype == "Float":
            pval = _num(pval)
        elif ptype == "Boolean":
            pval = _bool(pval)
        else:
            pval = _s(pval)
        entry = {"dataType": ptype}
        if pval is not None:
            entry["value"] = pval
        params[pname] = entry
        r += 1
    return name, parent, default_group, type_color, params


_PARAM_RE = re.compile(r"\{[^}]+\}")


def _maybe_param_binding(value):
    """If a value references a UDT parameter (e.g. '{Server}' or
    'ns=1;s=[{Device}]HRI352'), export it as a parameter binding object so
    Ignition resolves it per-instance. Otherwise return the plain string."""
    if value is None:
        return None
    if _PARAM_RE.search(value):
        return {"bindType": "parameter", "binding": value}
    return value


_BARE_PARAM_RE = re.compile(r"^\{[^}]+\}$")


def _alarm_value(v, kind="str"):
    """Resolve an alarm cell to a literal or a binding object, matching the
    formats Ignition exports:
      * whole cell '{Param}'      -> {"bindType": "UDTParameter", "value": "{Param}"}
      * 'expr:<text>' or {braces} -> {"bindType": "Expression",   "value": "<text>"}
      * anything else             -> literal (num / bool / str per `kind`)
    """
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        if s.lower().startswith("expr:"):
            return {"bindType": "Expression", "value": s[5:].strip()}
        if _BARE_PARAM_RE.match(s):
            return {"bindType": "UDTParameter", "value": s}
        if "{" in s and "}" in s:            # compound expr referencing params
            return {"bindType": "Expression", "value": s}
        if kind == "num":
            try:
                return float(s)
            except ValueError:
                return s
        if kind == "bool":
            return s.lower() in ("true", "1", "yes", "y")
        return s
    # non-string literal (number from Excel)
    if kind == "bool":
        return bool(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def _coerce_state_value(v):
    """States keep their type: TRUE/FALSE -> bool, whole number -> int, else text."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip()
    if s.lower() in ("true", "false"):
        return s.lower() == "true"
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s


def _read_states(path):
    """Return dict: tagName -> [ {label, value}, ... ]  for Metadata.states."""
    states_by_tag = {}
    try:
        df = pd.read_excel(path, sheet_name="States")
    except Exception:
        return states_by_tag
    for _, row in df.iterrows():
        tag = _s(row.get("tagName"))
        label = _s(row.get("label"))
        if not tag or label is None:
            continue
        states_by_tag.setdefault(tag, []).append(
            {"label": label, "value": _coerce_state_value(row.get("value"))})
    return states_by_tag


def _read_alarms(path):
    """Return dict: tagName -> [alarm dicts]."""
    alarms_by_tag = {}
    try:
        df = pd.read_excel(path, sheet_name="Alarms")
    except Exception:
        return alarms_by_tag
    for _, row in df.iterrows():
        tag = _s(row.get("tagName"))
        aname = _s(row.get("alarmName"))
        if not tag or not aname:
            continue
        alarm = {"name": aname}
        mode = _s(row.get("mode"))
        if mode:
            alarm["mode"] = mode
        # value columns that accept literal / {Param} / expr: bindings
        for key, col, kind in [
            ("setpointA", "setpointA", "num"),
            ("setpointB", "setpointB", "num"),
            ("label", "label", "str"),
            ("priority", "priority", "str"),
            ("enabled", "enabled", "bool"),
            ("timeOnDelaySeconds", "timeOnDelaySeconds", "num"),
            ("timeOffDelaySeconds", "timeOffDelaySeconds", "num"),
            ("displayPath", "displayPath", "str"),
        ]:
            val = _alarm_value(row.get(col), kind)
            if val is not None:
                alarm[key] = val
        ackn = _bool(row.get("ackNotesReqd"))
        if ackn is not None:
            alarm["ackNotesReqd"] = ackn
        alarms_by_tag.setdefault(tag, []).append(alarm)
    return alarms_by_tag


def _build_tag(row, alarms_by_tag, states_by_tag, default_group):
    name = _s(row.get("name"))
    if not name:
        return None
    tag = {}

    # Metadata custom property -- emitted on EVERY tag.
    metadata = {
        "shortDescription": _s(row.get("shortDescription")) or "",
        "longDescription": _s(row.get("longDescription")) or "",
    }
    if name in states_by_tag:
        metadata["states"] = states_by_tag[name]
    tag["Metadata"] = metadata

    value_source = _s(row.get("valueSource")) or "memory"
    tag["valueSource"] = value_source

    # source-specific fields
    if value_source == "opc":
        opc_path = _s(row.get("opcItemPath"))
        if opc_path:
            tag["opcItemPath"] = _maybe_param_binding(opc_path)
        opc_srv = _s(row.get("opcServer"))
        if opc_srv:
            tag["opcServer"] = _maybe_param_binding(opc_srv)
    elif value_source == "expr":
        expr = _s(row.get("expression"))
        if expr:
            tag["expression"] = expr

    tag["dataType"] = _s(row.get("dataType")) or "Boolean"

    doc = _s(row.get("documentation"))
    if doc:
        tag["documentation"] = doc
    tip = _s(row.get("tooltip"))
    if tip:
        tag["tooltip"] = tip

    tag["name"] = name
    tag["tagType"] = _s(row.get("tagType")) or "AtomicTag"

    eng_unit = _s(row.get("engUnit"))
    if eng_unit:
        tag["engUnit"] = eng_unit
    for jkey, col in [("engLow", "engLow"), ("engHigh", "engHigh")]:
        v = _num(row.get(col))
        if v is not None:
            tag[jkey] = v
    fmt = _s(row.get("formatString"))
    if fmt:
        tag["formatString"] = fmt

    group = _s(row.get("tagGroup")) or default_group
    if group:
        tag["tagGroup"] = group

    # value deadband
    db = _num(row.get("deadband"))
    if db is not None:
        tag["deadband"] = db
    dbm = _s(row.get("deadbandMode"))
    if dbm:
        tag["deadbandMode"] = dbm

    # history
    hist = _bool(row.get("historyEnabled"))
    if hist:
        tag["historyEnabled"] = True
        prov = _s(row.get("historyProvider"))
        if prov:
            tag["historyProvider"] = prov
        sm = _s(row.get("sampleMode"))
        if sm:
            tag["sampleMode"] = sm
        rate = _num(row.get("historySampleRate"), as_int=True)
        if rate is not None:
            tag["historySampleRate"] = rate
        units = _s(row.get("historySampleRateUnits"))
        if units:
            tag["historySampleRateUnits"] = units
        hdb = _num(row.get("historicalDeadband"))
        if hdb is not None:
            tag["historicalDeadband"] = hdb
        hstyle = _s(row.get("historicalDeadbandStyle"))
        if hstyle:
            tag["historicalDeadbandStyle"] = hstyle
        hmode = _s(row.get("historicalDeadbandMode"))
        if hmode:
            tag["historicalDeadbandMode"] = hmode

    # alarms attached from the Alarms sheet
    if name in alarms_by_tag:
        tag["alarms"] = alarms_by_tag[name]

    return tag


def build(input_file, output_file, udt_name_arg=None):
    # UDT-level info + parameters
    udt_name, parent, default_group, type_color, params = _read_udt_sheet(input_file)
    udt_name = udt_name_arg or udt_name

    # pick the tag sheet: prefer 'Tags', else first sheet (old flat format)
    xl = pd.ExcelFile(input_file)
    tag_sheet = "Tags" if "Tags" in xl.sheet_names else xl.sheet_names[0]
    df = pd.read_excel(input_file, sheet_name=tag_sheet).dropna(how="all")

    if not udt_name:
        udt_name = "UntitledUDT"
        print("[WARN] No UDT name found. Set it on the 'UDT' sheet or pass "
              "--udt-name. Using 'UntitledUDT'.")

    alarms_by_tag = _read_alarms(input_file)
    states_by_tag = _read_states(input_file)

    tags = []
    for _, row in df.iterrows():
        tag = _build_tag(row, alarms_by_tag, states_by_tag, default_group)
        if tag:
            tags.append(tag)

    result = {"name": udt_name, "tagType": "UdtType"}
    if parent:
        result["typeId"] = parent
    if params:
        result["parameters"] = params
    result["tags"] = tags
    if type_color is not None:
        result["typeColor"] = type_color

    # create the destination folder if it does not exist yet
    out_dir = os.path.dirname(os.path.abspath(output_file))
    os.makedirs(out_dir, exist_ok=True)

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f"[OK] Wrote {output_file}")
    print(f"     UDT Name : {udt_name}")
    if params:
        print(f"     Params   : {', '.join(params.keys())}")
    print(f"     Tags     : {len(tags)}")
    n_alarms = sum(len(a) for a in alarms_by_tag.values())
    if n_alarms:
        print(f"     Alarms   : {n_alarms}")
    n_states = sum(len(s) for s in states_by_tag.values())
    if n_states:
        print(f"     States   : {n_states} (across {len(states_by_tag)} tag(s))")
    return result


# ===========================================================================
def main():
    p = argparse.ArgumentParser(
        description="Generate an Ignition UDT template or build UDT JSON from it.")
    sub = p.add_subparsers(dest="cmd", required=True)

    pt = sub.add_parser("template", help="Create a fill-out Excel template with dropdowns.")
    pt.add_argument("output", nargs="?", default="Ignition_UDT_Template.xlsx")

    pb = sub.add_parser("build", help="Convert a filled workbook to Ignition UDT JSON.")
    pb.add_argument("input")
    pb.add_argument("output", nargs="?", default="output.json")
    pb.add_argument("--udt-name", default=None,
                    help="Override/supply the UDT name (for old flat sheets).")

    pi = sub.add_parser("import", help="Convert an Ignition UDT JSON back into a fillable workbook.")
    pi.add_argument("input", help="Path to the UDT .json file.")
    pi.add_argument("output", nargs="?", default="imported_template.xlsx",
                    help="Path to write the .xlsx (created/overwritten).")

    args = p.parse_args()
    if args.cmd == "template":
        make_template(args.output)
    elif args.cmd == "build":
        build(args.input, args.output, args.udt_name)
    elif args.cmd == "import":
        import_json(args.input, args.output)


if __name__ == "__main__":
    main()
